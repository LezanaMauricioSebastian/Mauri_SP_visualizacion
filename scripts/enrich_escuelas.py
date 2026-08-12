#!/usr/bin/env python3
"""Enrich GR/VA mesas-school GeoJSON with the national Padrón Oficial (CUE)."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

# Padrón Oficial 2025 (DIE-ICSE), row 13 = headers, data from row 14.
# Download: https://www.argentina.gob.ar/sites/default/files/2025.09.24_padron_oficial_establecimientos_educativos_die_icse_1.xlsx
PADRON_XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/padron-escuelas/padron2025.xlsx")

COL = {
    "jurisdiccion": 0,
    "sector": 1,
    "ambito": 2,
    "departamento": 3,
    "localidad": 5,
    "cueanexo": 7,
    "nombre": 8,
    "domicilio": 9,
    "cod_postal": 10,
    "telefono": 11,
    "email": 12,
    "modcomun": 13,
    "modespecial": 14,
    "modadultos": 15,
    "nvcjmaternal": 16,
    "nvcjinfantes": 17,
    "nvcprimario": 18,
    "nvcsecundario": 19,
    "nvcsecundinet": 20,
    "nvcsnu": 21,
    "nvcsnuinet": 22,
    "nvcsnucursos": 23,
    "nvceductempr": 24,
    "nvejinfantes": 25,
    "nveprimario": 26,
    "nvesecundario": 27,
    "nveadultos": 15,  # modalidad adultos flag; real adult levels below
    "nvaprimario": 29,
    "nvasecundario": 30,
    "nvaformprofes": 31,
    "nvaformprofesinet": 32,
    "nvaalfabetiz": 33,
    "nvhinicial": 34,
    "nvhprimario": 35,
    "nvhsecundario": 36,
    "tallerartist": 37,
    "servcomplem": 38,
}

FLAG_FIELDS = [
    "modcomun", "modespecial", "modadultos",
    "nvcjmaternal", "nvcjinfantes", "nvcprimario", "nvcsecundario", "nvcsecundinet",
    "nvcsnu", "nvcsnuinet", "nvcsnucursos",
    "nvceductempr", "nvejinfantes", "nveprimario", "nvesecundario",
    "nvaprimario", "nvasecundario", "nvaformprofes", "nvaformprofesinet", "nvaalfabetiz",
    "nvhinicial", "nvhprimario", "nvhsecundario",
    "tallerartist", "servcomplem",
]

TEXT_FIELDS = ["sector", "ambito", "domicilio", "cod_postal", "telefono", "email"]


def norm_cue(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return None
    return digits.zfill(9)


def flag(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return 1 if float(value) else 0
    except (TypeError, ValueError):
        return 1 if str(value).strip() not in {"0", "None", "none"} else 0


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_padron(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_cue = {}
    for i, row in enumerate(ws.iter_rows(min_row=14, values_only=True), start=14):
        cue = norm_cue(row[COL["cueanexo"]] if row else None)
        if not cue:
            continue
        rec = {
            "nombre_padron": clean_text(row[COL["nombre"]]),
            "departamento": clean_text(row[COL["departamento"]]),
            "localidad_padron": clean_text(row[COL["localidad"]]),
            "jurisdiccion": clean_text(row[COL["jurisdiccion"]]),
        }
        for key in TEXT_FIELDS:
            rec[key] = clean_text(row[COL[key]])
        for key in FLAG_FIELDS:
            rec[key] = flag(row[COL[key]])
        rec["nveadultos"] = rec["modadultos"]
        rec["fuente_padron"] = "Padrón Oficial de Establecimientos Educativos 2025"
        by_cue[cue] = rec
    wb.close()
    return by_cue


def enrich_feature(feature: dict, padron: dict) -> str:
    props = feature.setdefault("properties", {})
    cue = norm_cue(props.get("cueanexo"))
    if cue:
        props["cueanexo"] = cue
    rec = padron.get(cue) if cue else None
    if not rec:
        for key in FLAG_FIELDS + ["nveadultos"]:
            props.setdefault(key, 0)
        return "miss"

    for key, value in rec.items():
        if key == "nombre_padron":
            continue
        if key == "localidad_padron":
            if not props.get("localidad") and value:
                props["localidad"] = value
            continue
        if key in TEXT_FIELDS:
            if not props.get(key) and value:
                props[key] = value
            continue
        props[key] = value

    if rec.get("nombre_padron") and rec["nombre_padron"] != props.get("nombre"):
        props["nombre_padron"] = rec["nombre_padron"]
    return "hit"


def enrich_file(src: Path, dst: Path, padron: dict) -> dict:
    geo = json.loads(src.read_text(encoding="utf-8"))
    hits = misses = 0
    for feature in geo.get("features", []):
        status = enrich_feature(feature, padron)
        if status == "hit":
            hits += 1
        else:
            misses += 1
    dst.write_text(json.dumps(geo, ensure_ascii=False), encoding="utf-8")
    return {"src": src.name, "dst": dst.name, "total": hits + misses, "hits": hits, "misses": misses}


def main():
    if not PADRON_XLSX.exists():
        raise SystemExit(f"No se encontró el padrón: {PADRON_XLSX}")

    print(f"Leyendo padrón {PADRON_XLSX} ...")
    padron = load_padron(PADRON_XLSX)
    print(f"CUE en padrón: {len(padron)}")

    jobs = [
        (
            ROOT / "datos/gran_Resis/mesas_electores_x_escuelas_amgr.geojson",
            ROOT / "datos/gran_Resis/escuelas_amgr_enriquecido.geojson",
        ),
        (
            ROOT / "datos/villa_angela/mesas_electores_x_escuela_va.geojson",
            ROOT / "datos/villa_angela/escuelas_va_enriquecido.geojson",
        ),
    ]
    for src, dst in jobs:
        stats = enrich_file(src, dst, padron)
        print(
            f"{stats['src']} → {stats['dst']}: "
            f"{stats['hits']}/{stats['total']} con padrón "
            f"({stats['misses']} sin match)"
        )


if __name__ == "__main__":
    main()
